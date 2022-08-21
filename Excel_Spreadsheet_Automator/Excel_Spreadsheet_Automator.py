from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,4)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,5)
        corrected_price_cell.value = corrected_price
        
        
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
    chart.add_data(values)
    sheet.add_chart(chart,'A10')
        
    workbook.save(filename)

    
path = Path()
for  file in path.glob("*.xlsx"):
    process_workbook(file)